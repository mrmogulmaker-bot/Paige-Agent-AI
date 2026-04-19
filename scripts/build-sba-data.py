#!/usr/bin/env python3
"""
Refresh supabase/functions/_shared/sba-lender-data.ts from the SBA's public
Lender Activity Reports.

Run quarterly (or after each FYE update at data.sba.gov):
    python3 scripts/build-sba-data.py

The script downloads two XLSX files, extracts the lender × project-state
breakdown, keeps the top 30 SBA 7(a) lenders per state by approved dollars
(plus the full ~315-row 504 CDC list), and rewrites the TypeScript module.

Dependencies: openpyxl, requests
"""
import json
import os
import re
import sys
from collections import defaultdict
from pathlib import Path

import requests

try:
    import openpyxl
except ImportError:
    sys.exit("openpyxl not installed. Run: pip install openpyxl requests")

# Static SBA dataset resource IDs — these are stable across years; SBA replaces
# the file at the same resource_id each fiscal year.
SBA_7A_URL = (
    "https://data.sba.gov/dataset/b8c5c81b-fdfb-4bb0-9294-342fab141504/"
    "resource/514c761a-06d1-492e-a873-40ab693f7738/download"
)
SBA_504_URL = (
    "https://data.sba.gov/dataset/b8c5c81b-fdfb-4bb0-9294-342fab141504/"
    "resource/d4749532-03e7-4f0a-aa11-a2d80ed192ad/download"
)

OUT_PATH = Path(__file__).resolve().parent.parent / "supabase" / "functions" / "_shared" / "sba-lender-data.ts"


def download(url: str, dst: Path) -> None:
    print(f"Downloading {url}")
    r = requests.get(url, timeout=60)
    r.raise_for_status()
    dst.write_bytes(r.content)
    print(f"  -> {dst} ({dst.stat().st_size // 1024} KB)")


def parse_sheet(path: Path, sheet: str) -> list[dict]:
    wb = openpyxl.load_workbook(path, read_only=True)
    ws = wb[sheet]
    headers: list[str] | None = None
    out: list[dict] = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            headers = [str(c).strip() if c else f"col{idx}" for idx, c in enumerate(row)]
            continue
        if headers is None or all(c is None for c in row):
            continue
        out.append({headers[j]: row[j] for j in range(len(headers))})
    return out


def to_min(rows: list[dict], lender_key: str) -> list[dict]:
    return [
        {
            "l": r.get(lender_key),
            "lc": (r.get(f"{lender_key} City") or "").title(),
            "ls": r.get(f"{lender_key} State"),
            "ps": r.get("Project State"),
            "n": int(r.get("Approved Loans") or 0),
            "d": int(r.get("Approved Dollars") or 0),
        }
        for r in rows
        if r.get(lender_key) and r.get("Project State")
    ]


def top_per_state(rows: list[dict], n: int) -> list[dict]:
    bucket: dict[str, list[dict]] = defaultdict(list)
    for r in rows:
        bucket[r["ps"]].append(r)
    out: list[dict] = []
    for state_rows in bucket.values():
        state_rows.sort(key=lambda x: x["d"], reverse=True)
        out.extend(state_rows[:n])
    return out


def main() -> None:
    tmp = Path("/tmp/sba_refresh")
    tmp.mkdir(exist_ok=True)
    p7 = tmp / "sba_7a.xlsx"
    p5 = tmp / "sba_504.xlsx"

    download(SBA_7A_URL, p7)
    download(SBA_504_URL, p5)

    sba7a = to_min(parse_sheet(p7, "Lender_ProjSt"), "Lender")
    cdc = to_min(parse_sheet(p5, "CDC_ProjSt"), "CDC")

    sba7a_top = top_per_state(sba7a, 30)
    print(f"7(a) rows: {len(sba7a)} -> {len(sba7a_top)} after top-30/state")
    print(f"504 CDC rows: {len(cdc)} (full set kept)")

    # Try to read fiscal vintage from the About sheet
    try:
        wb = openpyxl.load_workbook(p7, read_only=True)
        about = wb["About"]
        vintage = "FY2025 (as of 9/30/2025)"
        for row in about.iter_rows(values_only=True):
            for c in row:
                if isinstance(c, str) and "as of" in c.lower():
                    vintage = c.strip()
                    break
    except Exception:
        vintage = "FY2025"

    body = f"""// AUTO-GENERATED from SBA FYE Lender Activity Reports (data.sba.gov)
// Source: https://data.sba.gov/dataset/lender-activity-reports
// 7(a): top 30 lenders per project state by approved dollars
// 504: full CDC × project-state breakdown
// Refresh quarterly via: python3 scripts/build-sba-data.py
// Field key: l=lender, lc=lender city, ls=lender HQ state, ps=project state, n=approved loans, d=approved dollars

export interface SbaLenderRow {{
  l: string; lc: string; ls: string; ps: string; n: number; d: number;
}}

export const SBA_DATA_VINTAGE = {json.dumps(vintage)};
export const SBA_DATA_SOURCE_URL = "https://data.sba.gov/dataset/lender-activity-reports";

export const SBA_7A_LENDERS_BY_STATE: SbaLenderRow[] = {json.dumps(sba7a_top, separators=(",", ":"))};

export const SBA_504_CDCS_BY_STATE: SbaLenderRow[] = {json.dumps(cdc, separators=(",", ":"))};
"""

    OUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    OUT_PATH.write_text(body)
    print(f"Wrote {OUT_PATH} ({OUT_PATH.stat().st_size // 1024} KB)")


if __name__ == "__main__":
    main()
